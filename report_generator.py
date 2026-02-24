from typing import Any, List
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
HEADER_FILL = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
PASS_FILL = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
FAIL_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
UNKNOWN_FILL = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
WRAP = Alignment(wrap_text=True, vertical="top")
HFONT = Font(bold=True)
def autosize(ws, min_w=12, max_w=70):
   for col in ws.columns:
       max_len = 0
       col_letter = get_column_letter(col[0].column)
       for cell in col:
           if cell.value is None:
               continue
           max_len = max(max_len, len(str(cell.value)))
       ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))
def style_header(ws, row=1):
   for cell in ws[row]:
       cell.font = HFONT
       cell.fill = HEADER_FILL
       cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
def add_table(ws, headers: List[str], rows: List[List[Any]], status_col_idx: int | None = None):
   ws.append(headers)
   style_header(ws, 1)
   for r in rows:
       ws.append(r)
   for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
       for cell in row:
           cell.alignment = WRAP
   if status_col_idx is not None:
       for r in range(2, ws.max_row + 1):
           st = ws.cell(r, status_col_idx).value
           st_u = str(st).upper()
           if st_u == "PASS":
               ws.cell(r, status_col_idx).fill = PASS_FILL
           elif st_u == "FAIL":
               ws.cell(r, status_col_idx).fill = FAIL_FILL
           else:
               ws.cell(r, status_col_idx).fill = UNKNOWN_FILL
   ws.freeze_panes = "A2"
   autosize(ws)
def build_excel_report(result) -> bytes:
   wb = Workbook()
   # ----------------------------
   # Dashboard
   # ----------------------------
   dash = wb.active
   dash.title = "Dashboard"
   dash["A1"] = "Firewall Governance Dashboard"
   dash["A2"] = f"Hostname: {result.meta.get('hostname')}"
   dash["A3"] = f"Platform: {result.meta.get('platform')}"
   dash["A4"] = f"Firmware: {result.meta.get('firmware_version')} (build {result.meta.get('firmware_build')})"
   bench = getattr(result, "benchmark_meta", {}) or {}
   dash["A5"] = f"Benchmark Pack: {bench.get('pack_name','')} {bench.get('pack_version','')}".strip()
   dash["A6"] = f"Benchmark Selection: {bench.get('selection','')}"
   cis_pass = sum(1 for c in result.cis if str(c["status"]).upper() == "PASS")
   cis_fail = sum(1 for c in result.cis if str(c["status"]).upper() == "FAIL")
   cis_unk  = sum(1 for c in result.cis if str(c["status"]).upper() not in ("PASS","FAIL"))
   total = len(result.cis) or 1
   compliance = round(100.0 * cis_pass / total, 2)
   scores = getattr(result, "scores", {}) or {}
   dash["A8"]  = "CIS Controls (subset)"; dash["B8"]  = len(result.cis)
   dash["A9"]  = "CIS PASS";             dash["B9"]  = cis_pass
   dash["A10"] = "CIS FAIL";             dash["B10"] = cis_fail
   dash["A11"] = "CIS UNKNOWN";          dash["B11"] = cis_unk
   dash["A12"] = "Compliance % (PASS/Total)"; dash["B12"] = compliance
   dash["A13"] = "Maturity % (weighted)";     dash["B13"] = scores.get("maturity_score", "")
   dash["A15"] = "Permissive Rules (>=MEDIUM)"; dash["B15"] = len(result.permissive)
   dash["A16"] = "Duplicates";                  dash["B16"] = len(result.duplicates)
   dash["A17"] = "Shadowed";                    dash["B17"] = len(result.shadowed)
   dash["A18"] = "Redundant";                   dash["B18"] = len(result.redundant)
   dash["A20"] = "Internet UTM coverage %";     dash["B20"] = result.sec_profile_coverage.get("utm_coverage_pct", 0)
   # Lifecycle on dashboard
   life = result.lifecycle_assessment or {}
   dash["A22"] = "Firmware Lifecycle Status"; dash["B22"] = life.get("firmware_status", "Review")
   dash["A23"] = "Security Exposure";         dash["B23"] = life.get("security_exposure", "Unknown")
   for r in range(8, 24):
       dash[f"A{r}"].font = HFONT
       dash[f"A{r}"].alignment = WRAP
   autosize(dash, min_w=18, max_w=80)
   # ----------------------------
   # CIS Scorecard
   # ----------------------------
   cis_ws = wb.create_sheet("CIS Scorecard")
   cis_headers = ["Control ID","Category","Control Name","Status","Observed","Expected","Weight","Remediation"]
   cis_rows = [
       [c["control_id"], c["category"], c["control_name"], c["status"], c["observed"], c["expected"], c["weight"], c["remediation"]]
       for c in result.cis
   ]
   add_table(cis_ws, cis_headers, cis_rows, status_col_idx=4)
   # ----------------------------
   # CIS Failures
   # ----------------------------
   fail_ws = wb.create_sheet("CIS Failures")
   fail_headers = ["Control ID","Category","Control Name","Status","Observed","Expected","Remediation"]
   fail_rows = [
       [c["control_id"], c["category"], c["control_name"], c["status"], c["observed"], c["expected"], c["remediation"]]
       for c in result.cis if str(c["status"]).upper() == "FAIL"
   ]
   if not fail_rows:
       fail_rows = [["-","-","No FAIL controls in current subset","-","-","-","-"]]
   add_table(fail_ws, fail_headers, fail_rows, status_col_idx=4)
   # ----------------------------
   # Lifecycle Risk Sheet
   # ----------------------------
   life_ws = wb.create_sheet("Lifecycle Risk")
   life_ws.append(["Attribute", "Value"])
   style_header(life_ws, 1)
   life_map = result.lifecycle_assessment or {}
   life_ws.append(["Platform", life_map.get("platform", "")])
   life_ws.append(["Platform Status", life_map.get("platform_status", "")])
   life_ws.append(["Firmware Version", life_map.get("firmware_version", "")])
   life_ws.append(["Firmware Build", life_map.get("firmware_build", "")])
   life_ws.append(["Firmware Status", life_map.get("firmware_status", "")])
   life_ws.append(["Security Exposure", life_map.get("security_exposure", "")])
   life_ws.append(["Recommendation", life_map.get("recommendation", "")])
   for row in life_ws.iter_rows(min_row=2, max_row=life_ws.max_row, min_col=1, max_col=2):
       for cell in row:
           cell.alignment = WRAP
   autosize(life_ws, min_w=22, max_w=80)
   # ----------------------------
   # Policies Raw
   # ----------------------------
   pol_ws = wb.create_sheet("Policies Raw")
   pol_headers = ["Policy ID","Name","Status","SrcIntf","DstIntf","SrcAddr","DstAddr","Service","Action","Schedule","Logtraffic","UTM Detected"]
   pol_rows = [
       [p["policy_id"], p["name"], p["status"], p["srcintf"], p["dstintf"], p["srcaddr"], p["dstaddr"],
        p["service"], p["action"], p["schedule"], p["logtraffic"], p["utm_detected"]]
       for p in result.policies_raw
   ]
   add_table(pol_ws, pol_headers, pol_rows)
   # ----------------------------
   # Permissive Rules
   # ----------------------------
   perm_ws = wb.create_sheet("Permissive Rules")
   perm_headers = ["Policy ID","Name","SrcIntf","DstIntf","SrcAddr","DstAddr","Service","Action","Logtraffic","UTM Detected","Risk Score","Severity","Reasons"]
   perm_rows = [
       [p["policy_id"], p["name"], p["srcintf"], p["dstintf"], p["srcaddr"], p["dstaddr"], p["service"],
        p["action"], p["logtraffic"], p["utm_detected"], p["risk_score"], p["severity"], p["reasons"]]
       for p in result.permissive
   ]
   add_table(perm_ws, perm_headers, perm_rows)
   # ----------------------------
   # Segmentation Matrix
   # ----------------------------
   seg_ws = wb.create_sheet("Network Segmentation")
   seg_headers = ["Source Interface","Destination Interface","Allowed Policy Count","Indicator"]
   seg_rows = [[s["srcintf"], s["dstintf"], s["policy_count"], s["indicator"]] for s in result.segmentation]
   add_table(seg_ws, seg_headers, seg_rows)
   # ----------------------------
   # Security Profile Coverage
   # ----------------------------
   cov_ws = wb.create_sheet("Security Profile Coverage")
   cov_ws.append(["Metric","Value"])
   style_header(cov_ws, 1)
   cov_ws.append(["Total policies", result.sec_profile_coverage["total_policies"]])
   cov_ws.append(["Internet-bound policies", result.sec_profile_coverage["internet_bound_policies"]])
   cov_ws.append(["Internet policies with UTM", result.sec_profile_coverage["internet_with_utm"]])
   cov_ws.append(["UTM coverage % (internet)", result.sec_profile_coverage["utm_coverage_pct"]])
   for row in cov_ws.iter_rows(min_row=2, max_row=cov_ws.max_row, min_col=1, max_col=2):
       for cell in row:
           cell.alignment = WRAP
   autosize(cov_ws, min_w=24, max_w=60)
   # ----------------------------
   # Lists
   # ----------------------------
   def list_sheet(name, headers, rows):
       ws = wb.create_sheet(name)
       add_table(ws, headers, rows)
       return ws
   list_sheet("Duplicate Rules", ["Policy ID","Duplicate Of","Criteria"],
              [[d["policy_id"], d["duplicate_of"], d["criteria"]] for d in result.duplicates])
   list_sheet("Shadowed Rules", ["Policy ID","Shadowed By","Reason"],
              [[s["policy_id"], s["shadowed_by"], s["reason"]] for s in result.shadowed])
   list_sheet("Redundant Rules", ["Policy ID","Covered By","Reason"],
              [[r["policy_id"], r["covered_by"], r["reason"]] for r in result.redundant])
   bio = BytesIO()
   wb.save(bio)
   return bio.getvalue()
